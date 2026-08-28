#!/usr/bin/env python3
"""
Step 1 of the lead recap/overview pipeline (see docs/Lead-Recap-Overview-Workflow.md).

Takes a local .xlsx export of the "Pipedrive Reengagement" Google Sheet (download it via
Google Drive's download_file_content with exportMimeType
application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, base64-decode, save to
disk -- there's no MCP tool that reads a live Sheet's tabs/cells directly) and produces:

  - leads-<tab>.json          one JSON array per source tab, each row a plain object plus
                               a "_key" field (the row's primary email, lowercased)
  - all-unique-emails.json    the union of every email across all tabs, for the SmartLead
                               bulk lookup step (lead-recap-smartlead-lookup.mjs)
  - batches/<tab>-NNN.json    each tab's rows chunked into BATCH_SIZE-row files
  - batch-manifest.json       [{tab, batchIndex, file (absolute path), count}, ...] -- the
                               `manifest` arg for lead-recap-workflow.mjs

Requires openpyxl (`pip install openpyxl`).

Usage: python3 lead-recap-prepare-batches.py <sheet-export.xlsx> <output-dir>
"""
import json
import os
import sys

import openpyxl

BATCH_SIZE = 15

# tab name in the workbook -> (output key, column holding the primary email)
TAB_MAP = {
    "Master List": ("master-list", "Email(s)"),
    "Smartlead Interested+followup": ("smartlead", "Email"),
    "Salesforce": ("salesforce", "Email"),
}


def first_email(v):
    if not v:
        return None
    for e in str(v).replace(";", ",").split(","):
        e = e.strip().lower()
        if "@" in e:
            return e
    return None


def rows_as_dicts(ws, headers):
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        d = {}
        for h, v in zip(headers, vals):
            if h is None or v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            d[h] = v
        if d:
            out.append(d)
    return out


def main(xlsx_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(os.path.join(out_dir, "batches"), exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_emails = set()
    manifest = []

    for sheet_name, (tab_key, email_col) in TAB_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found in workbook, skipping")
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        rows = rows_as_dicts(ws, headers)
        for r in rows:
            k = first_email(r.get(email_col))
            r["_key"] = k
            if k:
                all_emails.add(k)
        json.dump(rows, open(os.path.join(out_dir, f"leads-{tab_key}.json"), "w"))
        print(f"{sheet_name}: {len(rows)} rows")

        n_batches = (len(rows) + BATCH_SIZE - 1) // BATCH_SIZE
        for i in range(n_batches):
            chunk = rows[i * BATCH_SIZE : (i + 1) * BATCH_SIZE]
            fname = os.path.join(out_dir, "batches", f"{tab_key}-{i:03d}.json")
            json.dump(chunk, open(fname, "w"))
            manifest.append({"tab": tab_key, "batchIndex": i, "file": os.path.abspath(fname), "count": len(chunk)})

    json.dump(sorted(all_emails), open(os.path.join(out_dir, "all-unique-emails.json"), "w"))
    json.dump(manifest, open(os.path.join(out_dir, "batch-manifest.json"), "w"))
    print(f"\n{len(all_emails)} unique emails -> all-unique-emails.json")
    print(f"{len(manifest)} batches -> batch-manifest.json (feed this as the workflow's `manifest` arg)")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: lead-recap-prepare-batches.py <sheet-export.xlsx> <output-dir>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
