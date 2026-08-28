#!/usr/bin/env python3
"""
Step 3 (final) of the lead recap/overview pipeline (see
docs/Lead-Recap-Overview-Workflow.md). Merges the Workflow's per-batch recap results
back onto the original sheet rows and writes an output workbook with an added
Overview / Current Status / Last Touchpoint column per tab.

Matches POSITIONALLY within each (tab, batchIndex) batch -- i.e. recap[i] belongs to
batch_rows[i] -- rather than by the agent-echoed "key" string. This is deliberate:
spot-checking a real run showed agents preserve row order reliably even when they
occasionally botch the echoed key itself (a mangled/duplicate-flagged key, or the
literal string "null" for every row with no email on file -- both make key-based
matching ambiguous, but position doesn't). Falls back to key-based matching only if a
batch's recap count doesn't match its row count (logged loudly, never silent), and
marks anything still unmatched as "NOT PROCESSED" so gaps stay visible.

Requires openpyxl (`pip install openpyxl`).

Usage: python3 lead-recap-assemble-output.py <workflow-result.json> <batch-manifest.json> <output.xlsx>
  workflow-result.json -- the workflow's returned `{ results: [...] }` (the "result" field
                           of the Workflow tool's completion notification/output file)
  batch-manifest.json  -- the manifest written by lead-recap-prepare-batches.py
"""
import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Font

TAB_SHEET_NAMES = {
    "master-list": "Master List",
    "smartlead": "Smartlead Interested+followup",
    "salesforce": "Salesforce",
}


def first_email(v):
    if not v:
        return None
    for e in str(v).replace(";", ",").split(","):
        e = e.strip().lower()
        if "@" in e:
            return e
    return None


def main(workflow_result_path, manifest_path, out_path):
    wf = json.load(open(workflow_result_path))
    results = wf["results"] if "results" in wf else wf
    manifest = json.load(open(manifest_path))

    recap_for = {}  # (tab, batchIndex) -> ordered list of recap dicts
    key_fallback = {}  # (tab, key) -> recap, for the rare length-mismatch fallback
    for r in results:
        tab, bi = r["tab"], r["batchIndex"]
        recaps = r.get("recaps", [])
        recap_for[(tab, bi)] = recaps
        for rec in recaps:
            k = (rec.get("key") or "").strip().lower()
            if k and k != "null":
                key_fallback[(tab, k)] = rec

    batch_files_by_tab = {}
    for b in manifest:
        batch_files_by_tab.setdefault(b["tab"], []).append(b)

    mismatched_batches = []
    rows_by_tab = {}
    matched_recaps_by_tab = {}
    for tab, batches in batch_files_by_tab.items():
        rows, merged = [], []
        for b in sorted(batches, key=lambda x: x["batchIndex"]):
            batch_rows = json.load(open(b["file"]))
            rows.extend(batch_rows)
            recaps = recap_for.get((tab, b["batchIndex"]), [])
            if len(recaps) == len(batch_rows):
                merged.extend(recaps)
            else:
                mismatched_batches.append((tab, b["batchIndex"], len(batch_rows), len(recaps)))
                for row in batch_rows:
                    k = first_email(row.get("Email(s)") or row.get("Email"))
                    merged.append(key_fallback.get((tab, k)) if k else None)
        rows_by_tab[tab] = rows
        matched_recaps_by_tab[tab] = merged

    if mismatched_batches:
        print("WARNING -- length mismatch, used key-based fallback for these batches:")
        for tab, bi, nrows, nrecaps in mismatched_batches:
            print(f"  {tab} batch {bi}: {nrows} rows vs {nrecaps} recaps")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    summary = ["Tab,Rows,Recaps matched,Batches,Mismatched batches"]
    for tab, sheet_name in TAB_SHEET_NAMES.items():
        if tab not in rows_by_tab:
            continue
        rows, recaps = rows_by_tab[tab], matched_recaps_by_tab[tab]
        ws = wb.create_sheet(sheet_name)
        headers = [h for h in (list(rows[0].keys()) if rows else []) if h != "_key"]
        out_headers = headers + ["Overview", "Current Status", "Last Touchpoint"]
        ws.append(out_headers)
        for c in ws[1]:
            c.font = bold
        matched = 0
        for row, rec in zip(rows, recaps):
            vals = [row.get(h) for h in headers]
            if rec:
                vals += [rec.get("overview", ""), rec.get("current_status", ""), rec.get("last_touchpoint", "")]
                matched += 1
            else:
                vals += ["NOT PROCESSED — no recap returned for this lead", "", ""]
            ws.append(vals)
        for row_cells in ws.iter_rows(min_row=2):
            for c in row_cells[len(headers):]:
                c.alignment = wrap
        for i, h in enumerate(out_headers, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 60 if h == "Overview" else 22
        n_mismatch = len([1 for t, *_ in mismatched_batches if t == tab])
        print(f"{sheet_name}: {len(rows)} rows, {matched} recaps matched ({len(rows) - matched} NOT PROCESSED)")
        summary.append(f"{sheet_name},{len(rows)},{matched},{len(batch_files_by_tab[tab])},{n_mismatch}")

    ws = wb.create_sheet("Run Summary", 0)
    for line in summary:
        ws.append(line.split(","))
    for c in ws[1]:
        c.font = bold

    wb.save(out_path)
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: lead-recap-assemble-output.py <workflow-result.json> <batch-manifest.json> <output.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
